#!/usr/bin/env python3
"""Convert Shiller's ie_data.xls -> ie_data.xlsx for the Streamlit app.

The app (streamlit_app/app.py) reads `ie_data.xlsx` with openpyxl, which cannot
open the old binary .xls format that Shiller publishes. This script faithfully
converts the .xls (cached/computed values, all sheets) into the .xlsx the app
expects, preserving the exact "Data" sheet layout the app reads by column.

Usage
-----
    # 1. Download the latest ie_data.xls from
    #    https://shillerdata.com/  (or http://www.econ.yale.edu/~shiller/data.htm)
    #    and drop it in this folder (any of these names works):
    #        ie_data.xls   |   "ie_data (1).xls"   |   pass a path as arg 1
    #
    # 2. Run it (needs xlrd + openpyxl):
    python3 convert_data.py
    #    or point at a specific file:
    python3 convert_data.py ~/Downloads/ie_data.xls

    # 3. Commit + push so Streamlit Cloud redeploys:
    git add ie_data.xls ie_data.xlsx && git commit -m "Refresh Shiller data" && git push

Dependencies (one-time):
    pip install xlrd openpyxl
"""

import sys
from pathlib import Path

try:
    import xlrd  # reads the old .xls binary format
    import openpyxl  # writes the .xlsx the app reads
except ImportError:
    sys.exit(
        "Missing dependency. Install with:\n    pip install xlrd openpyxl\n"
        "(If your system blocks pip, use a venv:\n"
        "    python3 -m venv .venv && .venv/bin/pip install xlrd openpyxl\n"
        "    .venv/bin/python convert_data.py )"
    )

HERE = Path(__file__).resolve().parent
OUT = HERE / "ie_data.xlsx"

# Candidate source names, in priority order, unless one is passed on the CLI.
CANDIDATES = ["ie_data.xls", "ie_data (1).xls", "ie_databak.xls"]


def find_source() -> Path:
    if len(sys.argv) > 1:
        p = Path(sys.argv[1]).expanduser()
        if not p.exists():
            sys.exit(f"File not found: {p}")
        return p
    for name in CANDIDATES:
        p = HERE / name
        if p.exists():
            return p
    sys.exit(
        "No source .xls found. Put ie_data.xls in this folder, or pass a path:\n"
        "    python3 convert_data.py ~/Downloads/ie_data.xls"
    )


def convert(src: Path, out: Path) -> None:
    book = xlrd.open_workbook(str(src))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in book.sheet_names():
        s = book.sheet_by_name(name)
        ws = wb.create_sheet(title=name)
        for r in range(s.nrows):
            for c in range(s.ncols):
                v = s.cell_value(r, c)
                # Blank cells come back as "" from xlrd; write real empties so
                # the app's `if row[0] is None` / truthiness checks behave.
                ws.cell(row=r + 1, column=c + 1, value=(None if v == "" else v))
    wb.save(str(out))


def validate(out: Path) -> None:
    """Re-read with the app's exact logic and print the latest data point."""
    ws = openpyxl.load_workbook(str(out), data_only=True)["Data"]
    rows = [row for row in ws.iter_rows(min_row=9, values_only=True) if row[0] is not None]
    if not rows:
        sys.exit("Conversion produced no data rows — check the source file's 'Data' sheet.")
    last = rows[-1]
    print(f"  data rows : {len(rows)}")
    print(f"  last date : {last[0]}  (price {last[1]}, CPI {last[4]})")


def main() -> None:
    src = find_source()
    print(f"Converting {src.name}  ->  {OUT.name}")
    convert(src, OUT)
    validate(OUT)
    print(f"Wrote {OUT}")
    print("\nNext: git add ie_data.xls ie_data.xlsx && git commit -m 'Refresh Shiller data' && git push")


if __name__ == "__main__":
    main()
